from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import Car, RepairRecord, Part, StockPart
from .serializers import CarSerializer, RepairRecordSerializer, PartSerializer, StockPartSerializer
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import json

def login_view(request):
    """Страница входа"""
    if request.user.is_authenticated:
        return redirect('/')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('/')
        else:
            return render(request, 'login.html', {'error': 'Неверное имя пользователя или пароль'})
    return render(request, 'login.html')

def register_view(request):
    """Страница регистрации"""
    if request.user.is_authenticated:
        return redirect('/')
    
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('/')
    else:
        form = UserCreationForm()
    return render(request, 'register.html', {'form': form})

def logout_view(request):
    """Выход из системы"""
    logout(request)
    return redirect('login')

@login_required
def index_view(request):
    """Главная страница с гаражом"""
    return render(request, 'index.html')

@login_required
def car_detail_view(request, car_id):
    """Страница детального просмотра автомобиля с записями о ремонте"""
    try:
        car = Car.objects.get(pk=car_id, user=request.user)
    except Car.DoesNotExist:
        return redirect('/')
    return render(request, 'car_detail.html', {'car': car})

# API Views
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def car_list(request):
    """Список автомобилей пользователя"""
    if request.method == 'GET':
        cars = Car.objects.filter(user=request.user)
        serializer = CarSerializer(cars, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        try:
            print(f"POST request.data: {request.data}")
            print(f"POST request.user: {request.user}")
            serializer = CarSerializer(data=request.data)
            print(f"Serializer data: {serializer.initial_data}")
            if serializer.is_valid():
                car = serializer.save(user=request.user)
                print(f"Car saved: {car}")
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            print(f"Serializer errors: {serializer.errors}")
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            print(f"Error in car_list POST: {e}")
            print(error_trace)
            return Response({'error': str(e), 'detail': 'Внутренняя ошибка сервера', 'traceback': error_trace}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def car_detail(request, pk):
    """Детали, обновление, удаление автомобиля"""
    try:
        car = Car.objects.get(pk=pk, user=request.user)
    except Car.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = CarSerializer(car)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        try:
            serializer = CarSerializer(car, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    elif request.method == 'DELETE':
        car.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# Repair Records API
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def repair_record_list(request, car_id):
    """Список записей о ремонте для автомобиля"""
    try:
        car = Car.objects.get(pk=car_id, user=request.user)
    except Car.DoesNotExist:
        return Response({'error': 'Автомобиль не найден'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        records = RepairRecord.objects.filter(car=car)
        serializer = RepairRecordSerializer(records, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        try:
            data = request.data.copy()
            data['car'] = car.id
            # Получаем ID запчастей со склада (может быть список или отсутствовать)
            stock_part_ids = data.pop('stock_part_ids', [])
            if not isinstance(stock_part_ids, list):
                stock_part_ids = []
            
            serializer = RepairRecordSerializer(data=data)
            if serializer.is_valid():
                record = serializer.save(car=car)
                
                # Перемещаем запчасти со склада в ремонт
                if stock_part_ids:
                    stock_parts = StockPart.objects.filter(pk__in=stock_part_ids, car=car)
                    for stock_part in stock_parts:
                        # Создаем запчасть в ремонте из запчасти со склада
                        Part.objects.create(
                            repair_record=record,
                            name=stock_part.name,
                            part_code=stock_part.part_code,
                            manufacturer=stock_part.manufacturer,
                            quantity=stock_part.quantity,
                            cost=stock_part.cost
                        )
                        # Удаляем запчасть со склада
                        stock_part.delete()
                
                # Перезагружаем запись с запчастями
                serializer = RepairRecordSerializer(record)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            print(f"Error in repair_record_list POST: {e}")
            print(traceback.format_exc())
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def repair_record_detail(request, car_id, record_id):
    """Детали, обновление, удаление записи о ремонте"""
    try:
        car = Car.objects.get(pk=car_id, user=request.user)
        record = RepairRecord.objects.get(pk=record_id, car=car)
    except Car.DoesNotExist:
        return Response({'error': 'Автомобиль не найден'}, status=status.HTTP_404_NOT_FOUND)
    except RepairRecord.DoesNotExist:
        return Response({'error': 'Запись о ремонте не найдена'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = RepairRecordSerializer(record)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        try:
            data = request.data.copy()
            # Получаем ID запчастей со склада (может быть список или отсутствовать)
            stock_part_ids = data.pop('stock_part_ids', [])
            if not isinstance(stock_part_ids, list):
                stock_part_ids = []
            
            serializer = RepairRecordSerializer(record, data=data)
            if serializer.is_valid():
                serializer.save()
                
                # Перемещаем запчасти со склада в ремонт
                if stock_part_ids:
                    stock_parts = StockPart.objects.filter(pk__in=stock_part_ids, car=car)
                    for stock_part in stock_parts:
                        # Создаем запчасть в ремонте из запчасти со склада
                        Part.objects.create(
                            repair_record=record,
                            name=stock_part.name,
                            part_code=stock_part.part_code,
                            manufacturer=stock_part.manufacturer,
                            quantity=stock_part.quantity,
                            cost=stock_part.cost
                        )
                        # Удаляем запчасть со склада
                        stock_part.delete()
                
                # Перезагружаем запись с запчастями
                serializer = RepairRecordSerializer(record)
                return Response(serializer.data)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            print(f"Error in repair_record_detail PUT: {e}")
            print(traceback.format_exc())
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    elif request.method == 'DELETE':
        record.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# Parts API
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def part_create(request, car_id, record_id):
    """Создание запчасти для записи о ремонте"""
    try:
        car = Car.objects.get(pk=car_id, user=request.user)
        record = RepairRecord.objects.get(pk=record_id, car=car)
    except Car.DoesNotExist:
        return Response({'error': 'Автомобиль не найден'}, status=status.HTTP_404_NOT_FOUND)
    except RepairRecord.DoesNotExist:
        return Response({'error': 'Запись о ремонте не найдена'}, status=status.HTTP_404_NOT_FOUND)
    
    try:
        data = request.data.copy()
        serializer = PartSerializer(data=data)
        if serializer.is_valid():
            serializer.save(repair_record=record)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def part_detail(request, car_id, record_id, part_id):
    """Обновление и удаление запчасти"""
    try:
        car = Car.objects.get(pk=car_id, user=request.user)
        record = RepairRecord.objects.get(pk=record_id, car=car)
        part = Part.objects.get(pk=part_id, repair_record=record)
    except (Car.DoesNotExist, RepairRecord.DoesNotExist, Part.DoesNotExist):
        return Response({'error': 'Не найдено'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'PUT':
        serializer = PartSerializer(part, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        part.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# Stock Parts API
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def stock_part_list(request, car_id):
    """Список запчастей на складе для автомобиля"""
    try:
        car = Car.objects.get(pk=car_id, user=request.user)
    except Car.DoesNotExist:
        return Response({'error': 'Автомобиль не найден'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        stock_parts = StockPart.objects.filter(car=car)
        serializer = StockPartSerializer(stock_parts, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        try:
            data = request.data.copy()
            data['car'] = car.id
            serializer = StockPartSerializer(data=data)
            if serializer.is_valid():
                serializer.save(car=car)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            print(f"Error in stock_part_list POST: {e}")
            print(traceback.format_exc())
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def stock_part_detail(request, car_id, stock_part_id):
    """Детали, обновление, удаление запчасти на складе"""
    try:
        car = Car.objects.get(pk=car_id, user=request.user)
        stock_part = StockPart.objects.get(pk=stock_part_id, car=car)
    except Car.DoesNotExist:
        return Response({'error': 'Автомобиль не найден'}, status=status.HTTP_404_NOT_FOUND)
    except StockPart.DoesNotExist:
        return Response({'error': 'Запчасть на складе не найдена'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = StockPartSerializer(stock_part)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        serializer = StockPartSerializer(stock_part, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        stock_part.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_report_to_excel(request, car_id):
    """Экспорт отчета о ремонте в Excel"""
    try:
        car = Car.objects.get(pk=car_id, user=request.user)
    except Car.DoesNotExist:
        return Response({'error': 'Автомобиль не найден'}, status=status.HTTP_404_NOT_FOUND)
    
    # Get date range from query parameters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from or not date_to:
        return Response({'error': 'Необходимо указать даты начала и окончания периода'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Filter repairs by date range
    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    repairs = RepairRecord.objects.filter(
        car=car,
        date__gte=date_from_obj,
        date__lte=date_to_obj
    ).order_by('date')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет о ремонте"
    
    # Styles
    header_fill = PatternFill(start_color="00BFA5", end_color="00BFA5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Title
    ws['A1'] = f"Отчет о ремонте: {car.brand} {car.model}"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"VIN: {car.vin}"
    ws['A2'].font = Font(size=11)
    ws.merge_cells('A2:F2')
    
    ws['A3'] = f"Период: {date_from_obj.strftime('%d.%m.%Y')} - {date_to_obj.strftime('%d.%m.%Y')}"
    ws['A3'].font = Font(size=11)
    ws.merge_cells('A3:F3')
    
    # Headers
    headers = ['Дата', 'Пробег (км)', 'Выполненные работы', 'Стоимость работ (₽)', 'Запчасти', 'Стоимость запчастей (₽)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_style
        cell.alignment = center_alignment
    
    # Data
    row = 6
    total_work_cost = 0
    total_parts_cost = 0
    
    for repair in repairs:
        parts = Part.objects.filter(repair_record=repair)
        parts_cost = sum(float(part.cost) * (part.quantity or 1) for part in parts)
        parts_list = ', '.join([f"{part.name} ({part.part_code}) x{part.quantity or 1}" for part in parts])
        
        ws.cell(row=row, column=1, value=repair.date.strftime('%d.%m.%Y')).border = border_style
        ws.cell(row=row, column=2, value=repair.mileage).border = border_style
        ws.cell(row=row, column=3, value=repair.work_description).border = border_style
        ws.cell(row=row, column=4, value=float(repair.work_cost)).border = border_style
        ws.cell(row=row, column=5, value=parts_list if parts_list else '-').border = border_style
        ws.cell(row=row, column=6, value=float(parts_cost)).border = border_style
        
        # Format numbers
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        
        total_work_cost += float(repair.work_cost)
        total_parts_cost += parts_cost
        row += 1
    
    # Totals
    row += 1
    ws.cell(row=row, column=3, value="ИТОГО:").font = Font(bold=True, size=12)
    ws.cell(row=row, column=4, value=float(total_work_cost)).font = Font(bold=True)
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    ws.cell(row=row, column=6, value=float(total_parts_cost)).font = Font(bold=True)
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    
    row += 1
    ws.cell(row=row, column=3, value="ОБЩАЯ СТОИМОСТЬ:").font = Font(bold=True, size=12)
    ws.merge_cells(f'C{row}:D{row}')
    ws.cell(row=row, column=5, value=float(total_work_cost + total_parts_cost)).font = Font(bold=True, size=12)
    ws.cell(row=row, column=5).number_format = '#,##0.00'
    ws.merge_cells(f'E{row}:F{row}')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 20
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"report_{car.brand}_{car.model}_{date_from_obj.strftime('%Y%m%d')}_{date_to_obj.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
